from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class CX2PulseRecord:
    source_file: str
    timestamp: datetime
    capacity_ah: float
    rest_to_half_c_resistance_ohm: float
    half_c_to_one_c_resistance_ohm: float


REQUIRED_COLUMNS = [
    "Date_Time",
    "Step_Time(s)",
    "Step_Index",
    "Cycle_Index",
    "Current(A)",
    "Voltage(V)",
    "Discharge_Capacity(Ah)",
]
REQUIRED_PULSE_STEPS = (4, 5, 8, 9, 10)


def _group_rows_by_step(rows: list[dict]) -> dict[int, list[dict]]:
    by_step: dict[int, list[dict]] = {}
    for row in rows:
        try:
            step = int(row["Step_Index"])
        except (TypeError, ValueError):
            continue
        by_step.setdefault(step, []).append(row)
    return by_step


def _pulse_capacity(by_step: dict[int, list[dict]]) -> float:
    rest_before_discharge = by_step[4][-1]
    full_discharge_end = by_step[5][-1]
    return float(
        full_discharge_end["Discharge_Capacity(Ah)"]
        - rest_before_discharge["Discharge_Capacity(Ah)"]
    )


def _step_resistance(
    before: dict,
    after: dict,
    *,
    minimum_current_step_a: float = 0.1,
) -> float | None:
    delta_current = abs(float(after["Current(A)"]) - float(before["Current(A)"]))
    if delta_current < minimum_current_step_a:
        return None
    return (float(before["Voltage(V)"]) - float(after["Voltage(V)"])) / delta_current


def _pulse_measurements(
    by_step: dict[int, list[dict]],
) -> tuple[float, float, float, datetime] | None:
    pulse_rest = by_step[8][-1]
    half_c_first = by_step[9][0]
    half_c_last = by_step[9][-1]
    one_c_first = by_step[10][0]
    rest_resistance = _step_resistance(pulse_rest, half_c_first)
    incremental_resistance = _step_resistance(half_c_last, one_c_first)
    if rest_resistance is None or incremental_resistance is None:
        return None
    timestamp = half_c_first["Date_Time"]
    if not isinstance(timestamp, datetime):
        timestamp = pd.to_datetime(timestamp).to_pydatetime()
    return (
        _pulse_capacity(by_step),
        rest_resistance,
        incremental_resistance,
        timestamp,
    )


def _measurements_are_physical(
    capacity_ah: float,
    rest_resistance_ohm: float,
    incremental_resistance_ohm: float,
) -> bool:
    return bool(
        0.01 <= capacity_ah <= 2.0
        and 0.005 <= rest_resistance_ohm <= 1.0
        and 0.005 <= incremental_resistance_ohm <= 1.0
    )


def _process_cycle(rows: list[dict], source_file: str) -> CX2PulseRecord | None:
    by_step = _group_rows_by_step(rows)
    if not all(step in by_step for step in REQUIRED_PULSE_STEPS):
        return None
    measurements = _pulse_measurements(by_step)
    if measurements is None:
        return None
    capacity, rest_resistance, incremental_resistance, timestamp = measurements
    if not _measurements_are_physical(
        capacity,
        rest_resistance,
        incremental_resistance,
    ):
        return None
    return CX2PulseRecord(
        source_file,
        timestamp,
        capacity,
        rest_resistance,
        incremental_resistance,
    )


def _column_index(header) -> dict[str, int] | None:
    index = {str(name): i for i, name in enumerate(header) if name is not None}
    if any(column not in index for column in REQUIRED_COLUMNS):
        return None
    return index


def _row_dict(values, index: dict[str, int]) -> dict:
    return {column: values[index[column]] for column in REQUIRED_COLUMNS}


def _iter_cycle_rows(rows, index: dict[str, int]):
    current_cycle = None
    cycle_rows: list[dict] = []
    for values in rows:
        cycle = values[index["Cycle_Index"]]
        if cycle is None:
            continue
        if current_cycle is None:
            current_cycle = cycle
        if cycle != current_cycle:
            yield cycle_rows
            cycle_rows = []
            current_cycle = cycle
        cycle_rows.append(_row_dict(values, index))
    if cycle_rows:
        yield cycle_rows


def _collect_cycle_records(
    cycle_groups,
    source_file: str,
    max_cycles: int,
) -> list[CX2PulseRecord]:
    records: list[CX2PulseRecord] = []
    for cycle_rows in cycle_groups:
        record = _process_cycle(cycle_rows, source_file)
        if record is None:
            continue
        records.append(record)
        if len(records) >= max_cycles:
            break
    return records


def _records_from_xlsx(
    payload: bytes,
    source_file: str,
    max_cycles: int,
) -> list[CX2PulseRecord]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    try:
        sheets = [
            sheet for sheet in workbook.worksheets if sheet.title.lower() != "info"
        ]
        if not sheets:
            return []
        rows = sheets[0].iter_rows(values_only=True)
        index = _column_index(next(rows))
        if index is None:
            return []
        return _collect_cycle_records(
            _iter_cycle_rows(rows, index),
            source_file,
            max_cycles,
        )
    finally:
        workbook.close()


def extract_cx2_pulse_records(
    archive_path: str | Path,
    max_cycles_per_file: int = 25,
) -> pd.DataFrame:
    """Read representative complete pulse cycles from every XLSX export in the CX2-3 archive."""
    records: list[CX2PulseRecord] = []
    with zipfile.ZipFile(archive_path) as archive:
        names = sorted(
            name for name in archive.namelist() if name.lower().endswith(".xlsx")
        )
        for name in names:
            records.extend(
                _records_from_xlsx(
                    archive.read(name),
                    name,
                    max_cycles_per_file,
                )
            )
    if not records:
        raise ValueError("No complete CX2 pulse cycles were extracted")
    frame = pd.DataFrame([record.__dict__ for record in records])
    return (
        frame.sort_values("timestamp")
        .drop_duplicates(subset=["timestamp"])
        .reset_index(drop=True)
    )


def aggregate_cx2_by_file(records: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        records.groupby("source_file", as_index=False)
        .agg(
            timestamp=("timestamp", "median"),
            cycles_sampled=("capacity_ah", "size"),
            capacity_ah=("capacity_ah", "median"),
            pulse_resistance_ohm=("rest_to_half_c_resistance_ohm", "median"),
            incremental_resistance_ohm=(
                "half_c_to_one_c_resistance_ohm",
                "median",
            ),
        )
        .sort_values("timestamp")
        .reset_index(drop=True)
    )
    grouped["elapsed_days"] = (
        grouped["timestamp"] - grouped["timestamp"].iloc[0]
    ).dt.total_seconds() / 86400.0
    grouped["capacity_soh_pct"] = (
        100.0 * grouped["capacity_ah"] / grouped["capacity_ah"].iloc[0]
    )
    grouped["resistance_factor"] = (
        grouped["pulse_resistance_ohm"] / grouped["pulse_resistance_ohm"].iloc[0]
    )
    return grouped
